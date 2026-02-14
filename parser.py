from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import StrEnum

from httpx import AsyncClient, BasicAuth
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Optional

import xls_functions
import asyncio
import time
import constants
import platform
import colorama
import re
import os
from loguru import logger
from retry import retry
from messengers import send_service_tg_message


def set_work_dir():
    """Sets the working directory to the script's directory."""
    os.chdir(os.path.dirname(__file__))


def get_color(num: int) -> str:
    """Returns a color code for terminal output based on the input number."""
    return f'\033[{31 + num % 6}m'


class ComparisonField(StrEnum):
    SKU = 'art'
    NAME = 'name'


@dataclass
class Product:
    """Defines the structure of a product dictionary."""
    name: Optional[str]
    art: Optional[str]
    available: str
    link: str
    price: int = 0
    old_price: Optional[int] = None
    variant: Optional[str] = None


class Parser(ABC):
    """
    This class represents a web parser that extracts product information from a specific website.
    """

    # Configuration variables with descriptions
    number_of_workers = 3  # Number of concurrent asynchronous workers
    worker_timeout = 1  # Timeout in seconds for each worker
    worker_attempts = 2  # Number of attempts to fetch a page before giving up

    site = 'https://example.com'  # URL of the website to parse. To be overridden in subclasses
    price_file = 'example.xlsx'  # Name of the Excel file to store the results. To be overridden in subclasses
    output_path = constants.output_path  # Path to the directory where the Excel file will be saved
    use_discount = True  # Whether to consider supplier's discount in calculations
    max_products_per_page = ''  # String to append to category URLs to maximize product output

    excluded_links = []  # List of product links to exclude from parsing
    excluded_links_parts = []  # List of URL fragments to exclude links containing them
    excluded_categories_links_parts = []  # List of URL fragments to exclude categories containing them

    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36'
    headers = {'User-Agent': user_agent}
    extra_http_headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7,ru;q=0.6',
    }

    proxy_host: str = ''  # "your_proxy_host"
    proxy_port_http: int
    proxy_port_https: int
    proxy_username: str = ''
    proxy_password: str

    # Column numbers for storing data in the Excel file
    art_clmn = 1
    name_clmn = 2
    available_clmn = 3
    price_clmn = 4
    oldprice_clmn = 5
    group_clmn = 6
    link_clmn = 7
    variant_clmn = 8
    discount_clmn = 9
    present_at_site_clmn = 11
    unavailable_at_site_times_clmn = 12

    # Additional configuration options
    render_javascript = False  # Enable JavaScript rendering
    headless = True  # Run the browser in headless mode. False for clouflare protection
    use_connection_pool = True  # Reuse existing network connections if True, create new ones otherwise
    use_dalayed_availability = True  # Delay marking products as unavailable until multiple checks
    compared_product_field = 'art'  # Field used as primary key to identify products (art or name)
    max_unavailable_count = 4  # Maximum number of checks before marking a product as unavailable

    def __init__(self):
        set_work_dir()
        self._init_loggers()
        self._setup_proxies()
        self._setup_event_loop()
        self._init_workbook()

        self.client = None
        # Initialize the Playwright browser and context
        self.playwright = None
        self.browser = None
        self.context = None

        self.queue = asyncio.Queue()

    @abstractmethod
    async def get_categories_links(self, link: str) -> list[str]:
        """
        This method should be implemented in subclasses to retrieve a list of category links from the website.
        """
        pass

    @abstractmethod
    async def get_products_links(self, category_link: str) -> list[str]:
        """
        This method should be implemented in subclasses to extract product links from a given category page.
        """
        pass

    @abstractmethod
    async def get_product_info(self, product_link: str) -> list[Product]:
        """
        This method should be implemented in subclasses to parse product details from a product page and return a list
        of Product dictionaries.
        """
        pass

    def _init_loggers(self) -> None:
        """Initializes loggers for debug and info messages."""
        name = Path(self.price_file).stem
        logger.add(sink=f'./log/{name}_debug.log',
                   format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
                   level='DEBUG',
                   rotation='5 days',
                   retention='10 days',
                   )
        logger.add(sink=f'./log/{name}.log',
                   format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
                   level='INFO',
                   backtrace=True,
                   diagnose=True)
        logger.add(sink=lambda msg: send_service_tg_message(msg),
                   format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
                   level='ERROR',
                   backtrace=True,
                   diagnose=True)

    def _init_workbook(self) -> None:
        Path(constants.output_path).mkdir(parents=True, exist_ok=True)
        self.price_file_absolute = Path(constants.output_path) / self.price_file

        if self.use_dalayed_availability:
            self.wb: Workbook = xls_functions.init(self.price_file_absolute, create_on_error=True)
        else:
            self.wb: Workbook = Workbook()  # Create a new workbook
        self.sh: Worksheet = self.wb.active

        # Determine the column number for product comparison based on the configured field
        match self.compared_product_field:
            case 'art':
                self.compare_by_column_number = self.art_clmn
            case 'name':
                self.compare_by_column_number = self.name_clmn

        self.index = xls_functions.index_file(self.sh, self.compare_by_column_number)

    def _setup_proxies(self):
        self.proxies = None
        self.auth = None
        if self.proxy_host:
            self.proxies = {'http://': f'http://{self.proxy_host}:{self.proxy_port_http}',
                            'https://': f'http://{self.proxy_host}:{self.proxy_port_https}'}
            self.auth = BasicAuth(self.proxy_username, self.proxy_password) if self.proxy_username else None

        self.playwright_proxy = {
            'server': f'https://{self.proxy_host}:{self.proxy_port_https}',
            'username': self.proxy_username,
            'password': self.proxy_password
        } if self.proxy_host else None

    def _get_httpx_client(self) -> AsyncClient:
        return AsyncClient(follow_redirects=True,
                                  proxy=self.proxies,
                                  auth=self.auth,
                                  headers=self.headers,
                                  timeout=30)

    async def _init_playwright(self):
        """Initializes the Playwright library and creates a new browser context.
        for javascript rendering"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=self.headless)
        self.context = await self.browser.new_context(user_agent=self.user_agent,
                                                      extra_http_headers=self.extra_http_headers,
                                                      proxy=self.playwright_proxy)
        page = await self.context.new_page()
        await page.goto('about:blank')

    async def _close_playwright(self):
        await self.context.close()
        await self.browser.close()
        await self.playwright.stop()

    def _setup_event_loop(self):
        """Sets up the event loop policy for Windows systems."""
        if platform.system().lower() == 'windows' and not self.render_javascript:
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    async def get_html_page(self, url: str):
        """Fetches the HTML content of a page asynchronously."""
        if self.use_connection_pool:
            if self.client is None:
                self.client = self._get_httpx_client()
            r = await self.client.get(url=url)
        else:   # new client for every request
            async with self._get_httpx_client() as temp_client:
                r = await temp_client.get(url=url)
        return r.text

    async def get_javascript_page(self, url: str) -> str:
        if self.context is None:
            await self._init_playwright()
        page = await self.context.new_page()
        try:
            await page.goto(url)
            return await page.content()
        finally:
            await page.close()

    async def get_soup(self, url: str) -> BeautifulSoup:
        """Fetches the HTML content and parses it into a BeautifulSoup object."""
        html = await self.get_javascript_page(url) if self.render_javascript else await self.get_html_page(url)
        return BeautifulSoup(html, features='html.parser')

    async def _get_links_for_processing(self, site: str):
        """
        Collects unique product links from all categories and adds them to the queue for processing.
        """
        unique_links = set()
        category_links = await self.get_categories_links(site)
        for category_link in filter(self._is_valid_category_link, category_links):
            product_links = await self.get_products_links(category_link)
            for link in product_links:
                if link not in unique_links and self._is_valid_link(link):
                    unique_links.add(link)
                    await self.queue.put(link)
                    logger.debug(f'Appended link to queue: {link}')
        return len(unique_links)

    def _is_valid_link(self, link: str) -> bool:
        """Checks if the given link should be excluded based on configured URL fragments."""
        return link not in self.excluded_links and not any(part in link for part in self.excluded_links_parts)

    def _is_valid_category_link(self, link: str) -> bool:
        """Checks if the given link should be excluded based on configured URL fragments."""
        return link not in self.excluded_links and not any(part in link for part in self.excluded_categories_links_parts)

    def get_price(self, price_str: str) -> int:
        """Extracts the integer price from a string, handling commas and decimal points."""
        return int(''.join(re.findall(r'[\d,.]', price_str)).replace(',', '.').split('.')[0])

    def _write_product(self, product: Product):
        """Writes a product to the Excel sheet, determining the correct row based on the comparison field."""
        key = getattr(product, self.compared_product_field).lower().strip()
        row = self.index.get(key) or max(self.sh.max_row + 1, 2)
        self._write_product_to_xls(row=row, product=product)

    def _write_product_to_xls(self, row: int, product: Product):
        """Writes product information to the specified row in the Excel sheet."""
        self.sh.cell(row, self.name_clmn).value = product.name
        self.sh.cell(row, self.art_clmn).value = product.art
        self.sh.cell(row, self.price_clmn).value = product.old_price if product.old_price else product.price
        self.sh.cell(row, self.available_clmn).value = product.available
        self.sh.cell(row, self.link_clmn).value = product.link
        self.sh.cell(row, self.variant_clmn).value = product.variant
        if self.use_discount and product.old_price:
            self.sh.cell(row, self.discount_clmn).value = product.old_price - product.price
        else:
            self.sh.cell(row, self.discount_clmn).value = 0

        if self.use_dalayed_availability:
            self.sh.cell(row, self.present_at_site_clmn).value = 'present_at_site'

    def _process_unavailable(self):
        """
        Handles delayed availability logic, incrementing unavailable counts and marking products as unavailable after
        exceeding the threshold.
        """
        for i in range(2, self.sh.max_row + 1):
            if not self.sh.cell(i, self.present_at_site_clmn).value:
                if not self.sh.cell(i, self.unavailable_at_site_times_clmn).value:
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value = 1
                else:
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value += 1
                if self.sh.cell(i, self.unavailable_at_site_times_clmn).value >= self.max_unavailable_count:
                    self.sh.cell(i, self.available_clmn).value = '-'
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value = self.max_unavailable_count
            else:
                self.sh.cell(i, self.unavailable_at_site_times_clmn).value = None
                self.sh.cell(i, self.present_at_site_clmn).value = None

    @retry(max_tries=worker_attempts)
    async def get_product_info_advanced(self, product_link: str) -> list[Product]:
        return await self.get_product_info(product_link)

    async def worker(self, worker_id: int):
        """
        Worker function that processes product links from the queue, fetches product information, and writes it
        to the Excel sheet.
        """
        color = get_color(worker_id)
        logger.info(f'Worker {worker_id} - Starting parsing links of {self.site}')
        while True:
            product_link = await self.queue.get()
            try:
                logger.debug(color + f'Worker {worker_id}, ---- Getting {product_link}\n')
                products = await self.get_product_info_advanced(product_link)
                if not products:
                    logger.debug(f'***********  NO PRODUCTS FOUND ON THE PAGE  *********** {product_link}')
            except Exception as e:
                logger.error(f"ERROR {str(e)} parsing product {product_link} for {self.worker_attempts} attempts")
            else:
                for product in products:
                    logger.debug(color + f'Worker {worker_id} ---- Writing {product}\n')
                    self._write_product(product)

            self.queue.task_done()
            await asyncio.sleep(self.worker_timeout)

    async def main(self):
        """
        Main asynchronous function that manages the parsing process, creates workers, and handles the queue.
        """
        get_products_links_task = asyncio.create_task(self._get_links_for_processing(site=self.site))
        workers_tasks = [asyncio.create_task(self.worker(i)) for i in range(self.number_of_workers)]

        if not await get_products_links_task:
            for task in workers_tasks:
                task.cancel()
            raise Exception(f'Error parcing {self.site} Unable to get any product links')
        else:
            await self.queue.join()

        if self.render_javascript:
            await self._close_playwright()
        elif self.use_connection_pool:
            await self.client.aclose()

    def parse(self) -> None:
        """
        Starts the parsing process, measures execution time, and saves the results to the Excel file.
        """
        try:
            colorama.init()
            logger.info(f'Starting getting links for parsing for {self.site}')
            t0 = time.time()
            asyncio.run(self.main())
            t1 = time.time()
            logger.info(f'End parsing links of {self.site} Parsing Time = {t1 - t0:.02f} sec')

            if self.use_dalayed_availability:
                self._process_unavailable()

            self.wb.save(self.price_file_absolute)

        except Exception as e:
            logger.error(f'Error parsing {self.site} {str(e)}')
