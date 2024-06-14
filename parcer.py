from abc import ABC, abstractmethod
from httpx import AsyncClient, BasicAuth
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import TypedDict
import xls_functions
import asyncio
import time
import constants
import platform
import colorama
import re
import os
from loguru import logger


def set_work_dir():
    """Sets the working directory to the script's directory."""
    os.chdir(os.path.dirname(__file__))


def init_loggers(file_name: str) -> None:
    """Initializes loggers for debug and info messages."""
    set_work_dir()

    logger.add(sink='./log/' + file_name + '_debug.log',
               format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
               level='DEBUG',
               rotation='5 days',
               retention='10 days',
               )
    logger.add(sink='./log/' + file_name + '.log',
               format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
               level='INFO',
               backtrace=True,
               diagnose=True)
    logger.add(sink=lambda msg: constants.send_message(msg),
               format="{time:YYYY-MM-DD at HH:mm:ss} | {level} | {message}",
               level='ERROR',
               backtrace=True,
               diagnose=True)


def get_color(num: int) -> str:
    """Returns a color code for terminal output based on the input number."""
    return f'\033[{31 + num % 6}m'


class Product(TypedDict):
    """Defines the structure of a product dictionary."""
    name: str
    art: str | None
    price: int
    old_price: int | None
    available: str
    link: str
    variant: str | None


class Parcer(ABC):
    """
    This class represents a web parser that extracts product information from a specific website.
    """

    # Configuration variables with descriptions
    number_of_workers = 3  # Number of concurrent asynchronous workers
    worker_timeout = 1  # Timeout in seconds for each worker
    worker_attempts = 2  # Number of attempts to fetch a page before giving up

    site = ''  # URL of the website to parse
    price_file = ''  # Name of the Excel file to store the results
    use_discount = True  # Whether to consider supplier's discount in calculations
    max_products_per_page = ''  # String to append to category URLs to maximize product output

    excluded_links = []  # List of product links to exclude from parsing
    excluded_links_parts = []  # List of URL fragments to exclude links containing them

    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:77.0) Gecko/20100101 Firefox/77.0'}  # User-Agent header for HTTP requests
    proxy_host: str = ''  # "your_proxy_host"
    proxy_port_http: int
    proxy_port_https: int
    proxy_username: str = ''
    proxy_password: str

    # Column numbers for storing data in the Excel file
    art_clmn = 1
    name_clmn = 2
    available_clmn = 3
    price_clmn = 4
    oldprice_clmn = 5
    group_clmn = 6
    link_clmn = 7
    variant_clmn = 8
    discount_clmn = 9
    present_at_site_clmn = 11
    unavailable_at_site_times_clmn = 12

    # Additional configuration options
    use_connection_pool = True  # Reuse existing network connections if True, create new ones otherwise
    use_dalayed_availability = True  # Delay marking products as unavailable until multiple checks
    compared_product_field = 'art'  # Field used to compare and identify products (art or name)
    max_unavailable_count = 4  # Maximum number of checks before marking a product as unavailable

    def __init__(self):
        """Initializes the parser, sets up loggers, and prepares the Excel workbook."""
        init_loggers(self.price_file.removesuffix('.xlsx'))

        if self.proxy_host:
            self.proxies = {
                "http://": f"http://{self.proxy_host}:{self.proxy_port_http}",
                "https://": f"http://{self.proxy_host}:{self.proxy_port_https}",
            }
            self.auth = BasicAuth(self.proxy_username, self.proxy_password) if self.proxy_username else None
        else:
            self.proxies = None
            self.auth = None

        self.client = AsyncClient(follow_redirects=True, proxies=self.proxies, auth=self.auth)
        self.queue = asyncio.Queue()

        Path(constants.supl_path).mkdir(parents=True, exist_ok=True)
        self.price_file = Path(constants.supl_path) / self.price_file

        if self.use_dalayed_availability:
            self.wb: Workbook = xls_functions.init(self.price_file, create_on_error=True)
        else:
            self.wb: Workbook = Workbook()  # Create a new workbook
        self.sh: Worksheet = self.wb.active

        # Determine the column number for product comparison based on the configured field
        match self.compared_product_field:
            case 'art':
                self.compare_by_column_number = self.art_clmn
            case 'name':
                self.compare_by_column_number = self.name_clmn

        self.index = xls_functions.index_file(self.sh, self.compare_by_column_number)

        # Set event loop policy for Windows systems
        if platform.system().lower() == 'windows':
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    @abstractmethod
    async def get_categories_links(self, link: str) -> list[str]:
        """
        This method should be implemented in subclasses to retrieve a list of category links from the website.
        """
        pass

    @abstractmethod
    async def get_products_links(self, category_link: str) -> list[str]:
        """
        This method should be implemented in subclasses to extract product links from a given category page.
        """
        pass

    @abstractmethod
    async def get_product_info(self, product_link: str) -> list[Product] | None:
        """
        This method should be implemented in subclasses to parse product details from a product page and return a list of Product dictionaries.
        """
        pass

    async def get_page(self, url: str):
        """Fetches the HTML content of a page asynchronously."""
        if self.use_connection_pool:
            r = await self.client.get(url=url, headers=self.headers)
        else:
            # create and close a new client for each request to prohibit block by server for some sites
            temp_client = AsyncClient(follow_redirects=True, proxies=self.proxies, auth=self.auth)
            r = await temp_client.get(url=url, headers=self.headers)
            await temp_client.aclose()
        return r.text

    async def get_soup(self, url: str) -> BeautifulSoup:
        """Fetches the HTML content and parses it into a BeautifulSoup object."""
        text = await self.get_page(url)
        return BeautifulSoup(text, features='html.parser')

    async def get_links_for_processing(self, site: str):
        """
        Collects unique product links from all categories and adds them to the queue for processing.
        """
        check_duplicates = set()  # Store unique links
        for category_link in await self.get_categories_links(site):
            for product_link in await self.get_products_links(category_link + self.max_products_per_page):
                if product_link in check_duplicates:
                    continue
                if product_link in self.excluded_links:
                    continue
                if self.is_link_excluded_by_part(product_link):
                    continue

                await self.queue.put(product_link)
                check_duplicates.add(product_link)
                logger.debug(f'Appended link to queue: {product_link}')

        return len(check_duplicates)

    def get_price(self, price_str: str) -> int:
        """Extracts the integer price from a string, handling commas and decimal points."""
        return int(''.join(re.findall(r'[\d,.]', price_str)).replace(',', '.').split('.')[0])

    def is_link_excluded_by_part(self, checked_link: str) -> bool:
        """Checks if the given link should be excluded based on configured URL fragments."""
        for link_part in self.excluded_links_parts:
            if link_part in checked_link:
                return True
        return False

    def write_product_to_xls(self, row: int, product: Product):
        """Writes product information to the specified row in the Excel sheet."""
        self.sh.cell(row, self.art_clmn).value = product['art']
        self.sh.cell(row, self.name_clmn).value = product['name']
        self.sh.cell(row, self.price_clmn).value = product['old_price'] if product['old_price'] else product['price']
        self.sh.cell(row, self.available_clmn).value = product['available']
        self.sh.cell(row, self.link_clmn).value = product['link']
        self.sh.cell(row, self.variant_clmn).value = product['variant']
        if self.use_discount and product['old_price']:
            self.sh.cell(row, self.discount_clmn).value = product['old_price'] - product['price']
        else:
            self.sh.cell(row, self.discount_clmn).value = 0

        if self.use_dalayed_availability:
            self.sh.cell(row, self.present_at_site_clmn).value = 'present_at_site'

    def process_unavailable(self):
        """
        Handles delayed availability logic, incrementing unavailable counts and marking products as unavailable after exceeding the threshold.
        """
        for i in range(2, self.sh.max_row + 1):
            if not self.sh.cell(i, self.present_at_site_clmn).value:
                if not self.sh.cell(i, self.unavailable_at_site_times_clmn).value:
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value = 1
                else:
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value += 1
                if self.sh.cell(i, self.unavailable_at_site_times_clmn).value >= self.max_unavailable_count:
                    self.sh.cell(i, self.available_clmn).value = '-'
                    self.sh.cell(i, self.unavailable_at_site_times_clmn).value = self.max_unavailable_count
            else:
                self.sh.cell(i, self.unavailable_at_site_times_clmn).value = None
                self.sh.cell(i, self.present_at_site_clmn).value = None

    async def worker(self, i: int):
        """
        Worker function that processes product links from the queue, fetches product information, and writes it to the Excel sheet.
        """
        color = get_color(i)
        logger.info(f'Worker {i} - Starting parsing links of {self.site}')
        while True:
            product_link = await self.queue.get()
            for attempt in range(self.worker_attempts):
                logger.debug(color + f'Worker {i}, attempt {attempt} ---- Getting {product_link}\n')
                try:
                    some_products = await self.get_product_info(product_link)
                    if not some_products:
                        logger.debug(f'***********  NO PRODUCTS FOUND ON THE PAGE  *********** {product_link}')
                except Exception as e:
                    print(color + f'\n\n\n{str(e)}\n\n\n')
                    await asyncio.sleep(self.worker_timeout)
                    if attempt == self.worker_attempts - 1:
                        logger.error(f"Couldn't parce the link {product_link} for {self.worker_attempts} attempts")
                else:
                    for product in some_products:
                        logger.debug(color + f'Worker {i} ---- Writing {product}\n')
                        row = self.index.get(product[self.compared_product_field].lower().strip())
                        if row is None:
                            row = self.sh.max_row + 1
                            if row < 2:
                                row = 2
                        self.write_product_to_xls(row=row, product=product)
                    break  # Stop attempts on success

            await asyncio.sleep(self.worker_timeout)
            self.queue.task_done()

    async def main(self):
        """
        Main asynchronous function that manages the parsing process, creates workers, and handles the queue.
        """
        get_products_links_task = asyncio.create_task(self.get_links_for_processing(site=self.site))
        workers_tasks = [asyncio.create_task(self.worker(i)) for i in range(self.number_of_workers)]

        if not await get_products_links_task:
            logger.error(f"Error parcing {self.site}. Unable to get any product links")
            for task in workers_tasks:
                task.cancel()
        else:
            await self.queue.join()

        await self.client.aclose()

    @logger.catch
    def parse(self):
        """
        Starts the parsing process, measures execution time, and saves the results to the Excel file.
        """
        colorama.init()
        logger.info(f'Starting getting links for parsing for {self.site}')
        t0 = time.time()
        asyncio.run(self.main())
        t1 = time.time()
        logger.info(f'End parsing links of {self.site} Parsing Time = {t1-t0:.02f} sec')

        if self.use_dalayed_availability:
            self.process_unavailable()

        self.wb.save(self.price_file)
