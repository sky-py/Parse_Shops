import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from pathlib import Path


MAX_COLUMNS = 400
FIRST_DATA_ROW = 2

COLOR_YELLOW = 'F8F400'
COLOR_LIGHT_YELLOW = 'FAF894'
COLOR_RED = 'FF0000'
COLOR_LIGHT_BLUE = '90DCFE'
COLOR_BLUE = '538DD5'
COLOR_LIGHT_GREEN = 'A4F96B'
COLOR_GREEN = '5AD208'
COLOR_GREY = 'A6A6A6'
COLOR_LIGHT_GREY = 'D9D9D9'


def init(file: str | Path, create_on_error=False, **kwargs) -> openpyxl.Workbook:
    """
    Initializes an Excel workbook object.

    :param file: The path to the Excel file.
    :param create_on_error: If True, creates a new workbook if the file is not found.
    :return: The workbook object.
    :rtype: Workbook
    """
    if isinstance(file, str):
        file = Path(file)
    try:
        if file.suffix == '.xlsx':
            return openpyxl.load_workbook(file, **kwargs)
        else:
            from xls2xlsx import XLS2XLSX
            return XLS2XLSX(file).to_xlsx()
    except FileNotFoundError:
        if create_on_error:
            return openpyxl.Workbook()
        else:
            raise FileNotFoundError


def get_active_sheet(file: str | Path, create_on_error=False, **kwargs) -> Worksheet:
    """
    Gets the active sheet in an Excel workbook.

    :param file: The path to the Excel file.
    :param create_on_error: If True, creates a new workbook if the file is not found.
    :return: The active sheet object.
    :rtype: Worksheet
    """
    wb = init(file, create_on_error, **kwargs)
    return wb.active


def unmerge_all_cells(sh: Worksheet):
    for merge in list(sh.merged_cells):
        sh.unmerge_cells(range_string=str(merge))


def color_cell(cell: Cell, color: str = COLOR_YELLOW):
    """
    Sets the background color of a cell.

    :param cell: The cell object.
    :param color: The color code (e.g., 'FF0000' for red).
    """
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def color_row(row: int, color: str, sh: Worksheet):
    """
    Sets the background color of an entire row.

    :param row: The row number.
    :param color: The color code.
    :param sh: The worksheet object.
    """
    for column in range(1, MAX_COLUMNS + 1):
        color_cell(sh.cell(row, column), color)


def find_row(text: str, clmn_number: int, sh: Worksheet, first_data_row=FIRST_DATA_ROW) -> int:
    """
    Finds the row number that contains the specified text in the given column.

    :param text: The text to search for.
    :param clmn_number: The column number to search in.
    :param sh: The worksheet object.
    :return: The row number, or None if not found.
    :rtype: int
    """
    for row in range(first_data_row, sh.max_row + 1):
        if str(sh.cell(row, clmn_number).value).lower().strip() == str(text).lower().strip():
            return row


def find_column_by_name(clmn_name: str | list[str], sheet: Worksheet,
                        strip=True,
                        case_sensitive=False,
                        strict_match: bool = True) -> int:
    """
    Finds the column number by its name (header value).

    :param clmn_name: The column name.
    :param sheet: The worksheet object.
    :param strip: whether to strip whitespace from both column name and cell value
    :param case_sensitive: whether to perform a case-sensitive search
    :param strict_match: if False then just include
    :return: The column number, or None if not found.
    :rtype: int
    """
    if isinstance(clmn_name, str):
        clmn_name = [clmn_name]

    for column in range(1, MAX_COLUMNS + 1):
        cell_value = str(sheet.cell(1, column).value)
        if strip:
            cell_value = cell_value.strip()
        if not case_sensitive:
            cell_value = cell_value.lower()

        for clmn_text in clmn_name:
            if strip:
                clmn_text = clmn_text.strip()
            if not case_sensitive:
                clmn_text = clmn_text.lower()

            if strict_match:
                if cell_value == clmn_text:
                    return column
            elif clmn_text in cell_value:
                return column


def clear_row(row: int, sh: Worksheet):
    """
    Clears the contents of a row.

    :param row: The row number.
    :param sh: The worksheet object.
    """
    for column in range(1, MAX_COLUMNS + 1):
        sh.cell(row, column).value = None


def clear_row_from_column(row: int, column: int, sh: Worksheet):
    """
    Clears the contents of a row starting from the specified column.

    :param row: The row number.
    :param column: The starting column number.
    :param sh: The worksheet object.
    """
    for column in range(column, MAX_COLUMNS + 1):
        sh.cell(row, column).value = None


def index_file(sh: Worksheet, column: int, strip=True, lower=True, first_data_row=FIRST_DATA_ROW) -> dict:
    """
    Creates an index of the values in a column and their corresponding row numbers.

    :param sh: The worksheet object.
    :param column: The column number or name to index.
    :param strip: Whether to strip whitespace from cell values.
    :param lower: Whether to convert cell values to lowercase.
    :return: A dictionary where keys are cell values and values are row numbers.
    :rtype: dict
    """
    if isinstance(column, str):
        column = find_column_by_name(column, sh)

    index = dict()
    for row in range(first_data_row, sh.max_row + 1):
        if sh.cell(row, column).value:
            cell_value = str(sh.cell(row, column).value)
            if strip:
                cell_value = cell_value.strip()
            if lower:
                cell_value = cell_value.lower()
            index[cell_value] = row
    return index


def index_file_quick(sh: Worksheet, column: int | str, strip=True, lower=True, first_data_row=FIRST_DATA_ROW) -> dict:
    """
    Creates an index of the values in a column and their corresponding row numbers,
    but only includes the first occurrence of each value.

    :param sh: The worksheet object.
    :param column: The column number or name to index.
    :param strip: Whether to strip whitespace from cell values.
    :param lower: Whether to convert cell values to lowercase.
    :return: A dictionary where keys are cell values and values are row numbers.
    :rtype: dict
    """
    if isinstance(column, str):
        column = find_column_by_name(column, sh)

    index = dict()
    for row in range(first_data_row, sh.max_row + 1):
        if sh.cell(row, column).value:
            cell_value = str(sh.cell(row, column).value)
            if strip:
                cell_value = cell_value.strip()
            if lower:
                cell_value = cell_value.lower()
            if cell_value not in index:
                index[cell_value] = row
    return index


def get_row(index: dict, value, strip=True, lower=True) -> int | None:
    """
    Gets the row number corresponding to a value from an index.

    :param lower:
    :param strip:
    :param index: The index dictionary.
    :param value: The value to look up.
    :return: The row number, or None if not found.
    """
    if strip:
        value = value.strip()
    if lower:
        value = value.lower()

    return index.get(str(value))


def add_text_to_column(add_text, column, sh: Worksheet, first_data_row=FIRST_DATA_ROW, add_to_empty=True):
    """
    Adds text to the beginning or end of each cell in a column.

    :param add_text: The text to add.
    :param column: The column number.
    :param sh: The worksheet object.
    :param first_data_row: The starting row number.
    :param add_to_empty: Whether to add text to empty cells.
    """
    for row in range(first_data_row, sh.max_row + 1):
        if sh.cell(row, column).value is None or sh.cell(row, column).value == '':
            if not add_to_empty:
                continue
            else:
                cell_text = ''
        else:
            cell_text = str(sh.cell(row, column).value)
        sh.cell(row, column).value = cell_text + add_text


def copy_row(source_row: int, target_row: int, source_sheet: Worksheet, target_sheet: Worksheet=None, move=False):
    """
    Copies or moves a row from one location to another.

    :param source_row: The row number to copy from.
    :param target_row: The row number to copy to.
    :param source_sheet: The source worksheet.
    :param target_sheet: The target worksheet (optional, defaults to source sheet).
    :param move: If True, the source row is cleared after copying.
    """
    if target_sheet is None:
        target_sheet = source_sheet
    for column in range(1, MAX_COLUMNS + 1):
        target_sheet.cell(target_row, column).value = source_sheet.cell(source_row, column).value
        if move:
            clear_row(source_row, source_sheet)


def set_columns_width(width: int, sh: Worksheet) -> None:
    """
    Sets the width of all columns in a worksheet.

    :param width: The width to set.
    :param sh: The worksheet object.
    """
    for column_num in range(1, MAX_COLUMNS + 1):
        sh.column_dimensions[get_column_letter(column_num)].width = width


def copy_corresponding_column_values(*, column_out: int | str, sheet_out: Worksheet, index_column_out: int | str,
                                     column_in: int | str, sheet_in: Worksheet, index_column_in: int | str,
                                     first_data_row=FIRST_DATA_ROW):
    """
    Copies values from one column to another based on a corresponding index column.

    :param column_out: The column number or name in the output sheet.
    :param sheet_out: The output worksheet.
    :param index_column_out: The index column number or name in the output sheet.
    :param column_in: The column number or name in the input sheet.
    :param sheet_in: The input worksheet.
    :param index_column_in: The index column number or name in the input sheet.
    :param first_data_row: The starting row number for data (defaults to 2).
    """
    if isinstance(column_out, str):
        column_out = find_column_by_name(column_out, sheet_out)

    if isinstance(index_column_out, str):
        index_column_out = find_column_by_name(index_column_out, sheet_out)

    if isinstance(column_in, str):
        column_in = find_column_by_name(column_in, sheet_in)

    if isinstance(index_column_in, str):
        index_column_in = find_column_by_name(index_column_in, sheet_in)

    index = index_file_quick(sheet_in, index_column_in)
    for i in range(first_data_row, sheet_out.max_row + 1):
        index_value = str(sheet_out.cell(i, index_column_out).value)
        source_row = index.get(index_value.lower().strip())
        if source_row is not None:
            new_value = sheet_in.cell(source_row, column_in).value
            sheet_out.cell(i, column_out).value = new_value


def make_differences_workbook(*, sheet_out: Worksheet, index_column_out: int | str,
                              sheet_in: Worksheet, index_column_in: int | str,
                              first_data_row=FIRST_DATA_ROW) -> Workbook:
    """
    Creates a new workbook containing the rows that are in sheet_out but not in sheet_in,
    based on a comparison of index columns.

    :param sheet_out: The output worksheet.
    :param index_column_out: The index column number or name in the output sheet.
    :param sheet_in: The input worksheet.
    :param index_column_in: The index column number or name in the input sheet.
    :param first_data_row: The starting row number for data (defaults to 2).
    :return: The new workbook object.
    :rtype: Workbook
    """
    if isinstance(index_column_out, str):
        index_column_out = find_column_by_name(index_column_out, sheet_out)
    if isinstance(index_column_in, str):
        index_column_in = find_column_by_name(index_column_in, sheet_in)

    wb_diff = openpyxl.Workbook()
    sh_diff = wb_diff.active
    copy_row(source_row=1, source_sheet=sheet_out, target_row=1, target_sheet=sh_diff)
    pos = first_data_row
    index = index_file_quick(sheet_in, index_column_in)
    for i in range(first_data_row, sheet_out.max_row + 1):
        index_value = str(sheet_out.cell(i, index_column_out).value)
        source_row = index.get(index_value.lower().strip())
        if source_row is None:
            copy_row(source_row=i, source_sheet=sheet_out, target_row=pos, target_sheet=sh_diff)
            pos += 1
    return wb_diff


def get_rows_by_key(*, search_phrase: str, search_column: int | str, sh: Worksheet, first_data_row=FIRST_DATA_ROW) -> list[int]:
    """
    Retrieves a list of row numbers where a specific value is found in a given column.

    :param search_phrase: The value to search for.
    :param search_column: The column number or name to search in.
    :param sh: The worksheet object to search within.
    :param first_data_row: The starting row number for data (default is 2).
    :return: A list of row numbers where the search_phrase is found.
    :rtype: list[int]
    """
    rows = []
    if isinstance(search_column, str):
        search_column = find_column_by_name(search_column, sh)
    for i in range(first_data_row, sh.max_row + 1):
        if sh.cell(i, search_column).value:
            if str(sh.cell(i, search_column).value).lower().strip() == search_phrase.lower().strip():
                rows.append(i)
    return rows


def get_column_values(*, source: Path | str | Worksheet, column: str | int, exclude_empty: bool = False,
                      unique: bool = False, lower: bool = False, strip: bool = False, first_data_row=FIRST_DATA_ROW) -> list:
    """
    Extracts values from a specific column in a worksheet, with options for filtering and formatting.

    :param source: The path to the Excel file, or a Worksheet object.
    :param column: The column number or name to extract values from.
    :param exclude_empty: If True, excludes empty cells from the result (default is False).
    :param unique: If True, returns only unique values (default is False).
    :param lower: If True, converts all values to lowercase (default is False).
    :param strip: If True, removes leading and trailing whitespace from values (default is False).
    :return: A list of values extracted from the specified column.
    :rtype: list
    """
    if not isinstance(source, Worksheet):
        source = init(source).active

    if isinstance(column, str):
        column = find_column_by_name(column, source)

    keywords = []
    for i in range(first_data_row, source.max_row + 1):
        if exclude_empty and not source.cell(i, column).value:
            continue
        value = source.cell(i, column).value
        if strip:
            value = value.strip()
        if lower:
            value = value.lower()
        keywords.append(value)
    return list(set(keywords)) if unique else keywords
